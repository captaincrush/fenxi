# config.py
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------- 配置 ----------
CURRENT_DIR = os.getcwd()

# 销量分析表位置配置
SALES_TABLE_START_COL = 13  # T列
SALES_TABLE_START_ROW = 3   # 第3行

# 转化率分析表位置配置  
CONV_TABLE_START_COL = 13   # T列
CONV_TABLE_START_ROW = 12   # 第12行

# ACOS分析表格配置
ACOS_ANALYSIS_START_COL = 13  # W列开始
ACOS_ANALYSIS_START_ROW = 18   # 第3行开始

# 表格标题样式
SALES_TITLE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 销量标题深蓝色
CONV_TITLE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # 转化率标题深红色
HEADER_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")      # 表头浅紫色

# 五列四行汇总表起点
SUMMARY_START_COL = 21  # M
SUMMARY_START_ROW = 4   # 第3行

# 样式配置
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BOLD_FONT = Font(bold=True)