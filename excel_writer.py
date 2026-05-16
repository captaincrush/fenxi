# excel_writer.py
import openpyxl
from openpyxl.comments import Comment
from copy import copy
from config import *
from utils import *

class ExcelWriter:
    def __init__(self, ws_out):
        self.ws_out = ws_out
        self.sku_row_map = {}
        self.sku_col_idx = None

    def adjust_original_data_simple(self):
        """
        调整原始数据：第一列列宽增加到25个字符，按Units Ordered降序排列
        基于Excel中现有的数据
        """
        # 1. 调整第一列列宽
        first_column_letter = openpyxl.utils.get_column_letter(1)  # A列
        self.ws_out.column_dimensions[first_column_letter].width = 25
        
        # 2. 按Units Ordered降序排列
        self._sort_by_units_ordered_simple()

    def _sort_by_units_ordered_simple(self):
        """
        基于Excel中现有数据按Units Ordered降序排列
        """
        def find_col_by_keywords(keywords):
            for idx, cell in enumerate(self.ws_out[1], start=1):
                if cell.value and any(kw.lower() in str(cell.value).lower() for kw in keywords):
                    return idx
            return None

        units_col_idx = find_col_by_keywords(["units ordered", "unitsordered"])
        if not units_col_idx:
            return

        sku_col_idx = find_col_by_keywords(["sku"])
        raw_max_col = 0
        for idx, cell in enumerate(self.ws_out[1], start=1):
            if cell.value is not None:
                raw_max_col = idx
        if raw_max_col == 0:
            return

        def snapshot_cell(cell):
            return {
                "value": cell.value,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "comment": copy(cell.comment) if cell.comment else None,
                "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
            }

        def restore_cell(cell, snapshot):
            cell.value = snapshot["value"]
            cell.font = snapshot["font"]
            cell.fill = snapshot["fill"]
            cell.border = snapshot["border"]
            cell.alignment = snapshot["alignment"]
            cell.number_format = snapshot["number_format"]
            cell.protection = snapshot["protection"]
            cell.comment = snapshot["comment"]
            cell.hyperlink = snapshot["hyperlink"]
        
        # 获取所有数据行（从第2行开始）
        data_rows = []
        data_row_numbers = []
        
        # 读取数据行（从第2行开始）
        for row in range(2, self.ws_out.max_row + 1):
            if sku_col_idx:
                sku_value = self.ws_out.cell(row=row, column=sku_col_idx).value
                if sku_value is None or not str(sku_value).strip():
                    continue

            row_data = []
            for col in range(1, raw_max_col + 1):
                row_data.append(snapshot_cell(self.ws_out.cell(row=row, column=col)))
            
            # 获取Units Ordered值
            units_value = 0
            if units_col_idx - 1 < len(row_data):
                try:
                    units_cell_value = row_data[units_col_idx - 1]["value"]
                    units_value = float(units_cell_value) if units_cell_value else 0
                except (ValueError, TypeError):
                    units_value = 0
            
            data_rows.append({
                'row_index': row,
                'data': row_data,
                'units': units_value
            })
            data_row_numbers.append(row)
        
        # 按Units Ordered降序排序
        data_rows.sort(key=lambda x: x['units'], reverse=True)
        
        # 重新写入排序后的数据
        for i, row_info in zip(data_row_numbers, data_rows):
            row_data = row_info['data']
            for col in range(1, raw_max_col + 1):
                if col - 1 < len(row_data):
                    restore_cell(self.ws_out.cell(row=i, column=col), row_data[col - 1])

    def _clear_range(self, start_row, start_col, end_row, end_col):
        """清空区域内容，并解除与区域重叠的合并单元格。"""
        for merged_range in list(self.ws_out.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            overlaps = not (
                max_row < start_row or min_row > end_row or
                max_col < start_col or min_col > end_col
            )
            if overlaps:
                self.ws_out.unmerge_cells(str(merged_range))

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                self.ws_out.cell(row=r, column=c).value = None

    def find_sku_column(self):
        """找 SKU 列的位置"""
        for idx, cell in enumerate(self.ws_out[1], start=1):
            if cell.value and "sku" in str(cell.value).lower():
                self.sku_col_idx = idx
                break
        if not self.sku_col_idx:
            raise ValueError("在最新表格中未能找到 SKU 列（第一行包含 'SKU' 字样）")
            
    def build_sku_row_map(self):
        """构建 SKU -> 行 映射"""
        for r in range(2, self.ws_out.max_row + 1):
            v = self.ws_out.cell(row=r, column=self.sku_col_idx).value
            if v is not None:
                self.sku_row_map[str(v)] = r

    def write_acos_column(self, data_latest):
        """写入ACOS列"""
        col_idx_acos = self.sku_col_idx + 10
        
        # 设置标题
        title_cell = self.ws_out.cell(row=1, column=col_idx_acos)
        title_cell.value = "AD-ACOS(%)"
        title_cell.font = BOLD_FONT

        # 写入数据行
        for sku, row in self.sku_row_map.items():
            acos_value = data_latest.get(sku, {}).get("acos")
            if acos_value is not None:
                cell = self.ws_out.cell(row=row, column=col_idx_acos, value=f"{acos_value:.2f}")
                cell.number_format = '0.00'
                cell.value = float(cell.value)

    #新增
    def write_acos_analysis_table(self, top_acos_results):
        """
        写入ACOS分析表格：显示本周ACOS>10%且销量增长最快的SKU
        top_acos_results: [{'sku': 'xxx', 'acos1': 0.1, 'acos2': 0.15, 'sales_growth_rate': 50.0, 'u1': 100, 'u2': 150}]
        """
        start_col = 13  # W列开始
        start_row = 21   # 第3行开始
        
        # 清空区域
        self._clear_range(start_row, start_col, start_row + 9, start_col + 6)
        
        # 大标题行 - "高ACOS高增长SKU"
        title_cell = self.ws_out.cell(row=start_row, column=start_col, value="本周广告ACOS>10%，销量增长最快的3个SKU")
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = THIN_BORDER
        self.ws_out.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 6)
        
        # 表头行
        header_row = start_row + 1
        headers = ["排名", "SKU", "上周销量", "本周销量", "销量增长", "本周ACOS", "本周广告出单占比"]
        for col_idx, header in enumerate(headers):
            cell = self.ws_out.cell(row=header_row, column=start_col + col_idx, value=header)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT

        # 写入数据
        current_row = header_row + 1
        for i, result in enumerate(top_acos_results, 1):
            # 排名
            cell = self.ws_out.cell(row=current_row, column=start_col, value=f"Top{i}")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = RED_FILL
            
            # SKU
            cell = self.ws_out.cell(row=current_row, column=start_col + 1, value=result['sku'])
            cell.border = THIN_BORDER
            # cell.fill = RED_FILL
            
            # 上周销量
            cell = self.ws_out.cell(row=current_row, column=start_col + 2, value=safe_round_units(result['u1']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = RED_FILL
            
            # 本周销量
            cell = self.ws_out.cell(row=current_row, column=start_col + 3, value=safe_round_units(result['u2']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = RED_FILL
            
            # 销量增长
            growth_cell = self.ws_out.cell(row=current_row, column=start_col + 4, 
                                        value=format_percent_with_sign(result['sales_growth_rate']))
            growth_cell.border = THIN_BORDER
            growth_cell.alignment = CENTER_ALIGNMENT
            # growth_cell.fill = RED_FILL
            
            # 本周ACOS
            acos_cell = self.ws_out.cell(row=current_row, column=start_col + 5, 
                                    value=f"{result['acos2']:.2f}%")
            acos_cell.border = THIN_BORDER
            acos_cell.alignment = CENTER_ALIGNMENT
            # acos_cell.fill = RED_FILL

            # 广告出单占比
            ratio_cell = self.ws_out.cell(row=current_row, column=start_col + 6, 
                                        value=f"{result['ad_sales_ratio2']:.2f}%")
            ratio_cell.border = THIN_BORDER
            ratio_cell.alignment = CENTER_ALIGNMENT
            # ratio_cell.fill = RED_FILL
            
            current_row += 1
    
    #新增
    def write_high_conversion_table(self, high_conversion_results):
        """
        写入高转化率产品表格：7天订单数量大于10,转化率最高的3个产品
        """
        start_col = 13  # W列开始
        start_row = 29  # 第15行开始（放在ACOS表格下面）
        
        # 清空区域
        self._clear_range(start_row, start_col, start_row + 9, start_col + 5)
        
        # 大标题行
        title_cell = self.ws_out.cell(row=start_row, column=start_col, value="7天订单数量>10，转化率最高的3个SKU")
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = THIN_BORDER
        self.ws_out.merge_cells(start_row=start_row, start_column=start_col, 
                            end_row=start_row, end_column=start_col + 4)
        
        # 表头行
        header_row = start_row + 1
        headers = ["排名", "SKU", "订单量", "点击量", "转化率"]
        for col_idx, header in enumerate(headers):
            cell = self.ws_out.cell(row=header_row, column=start_col + col_idx, value=header)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT

        # 写入数据
        current_row = header_row + 1
        for i, result in enumerate(high_conversion_results, 1):
            # 排名
            cell = self.ws_out.cell(row=current_row, column=start_col, value=f"Top{i}")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = GREEN_FILL  # 用绿色表示优质产品
            
            # SKU
            cell = self.ws_out.cell(row=current_row, column=start_col + 1, value=result['sku'])
            cell.border = THIN_BORDER
            # cell.fill = GREEN_FILL
            
            # 订单数量
            cell = self.ws_out.cell(row=current_row, column=start_col + 2, value=safe_round_units(result['units']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = GREEN_FILL
            
            # 会话数量
            cell = self.ws_out.cell(row=current_row, column=start_col + 3, value=safe_round_units(result['sessions']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = GREEN_FILL
            
            # 转化率
            cell = self.ws_out.cell(row=current_row, column=start_col + 4, value=f"{result['conversion_rate']:.2f}%")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = GREEN_FILL
            
            current_row += 1

    #新增
    def write_efficient_products_table(self, efficient_results):
        """
        写入高效产品表格：7天订单数量大于10,转化率>10%,ACOS<5%
        """
        start_col = 21  # S列开始（19对应S列，28才对应AB列）
        start_row = 12  # 第12行开始
        
        # 清空区域
        self._clear_range(start_row, start_col, start_row + 9, start_col + 6)
        
        # 大标题行
        title_cell = self.ws_out.cell(row=start_row, column=start_col, value="7天订单数量>10，转化率>10%，且综合ACOS<5%的SKU")
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = THIN_BORDER
        self.ws_out.merge_cells(start_row=start_row, start_column=start_col, 
                            end_row=start_row, end_column=start_col + 6)
        
        # 表头行
        header_row = start_row + 1
        headers = ["SKU", "订单量", "点击量", "转化率", "ACOS", "广告花费", "销售额"]
        for col_idx, header in enumerate(headers):
            cell = self.ws_out.cell(row=header_row, column=start_col + col_idx, value=header)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT

        # 写入数据
        current_row = header_row + 1
        for result in efficient_results:
            # SKU
            cell = self.ws_out.cell(row=current_row, column=start_col, value=result['sku'])
            cell.border = THIN_BORDER
            # cell.fill = BLUE_FILL  # 用蓝色表示高效产品
            
            # 订单数量
            cell = self.ws_out.cell(row=current_row, column=start_col + 1, value=safe_round_units(result['units']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            # 会话数量
            cell = self.ws_out.cell(row=current_row, column=start_col + 2, value=safe_round_units(result['sessions']))
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            # 转化率
            cell = self.ws_out.cell(row=current_row, column=start_col + 3, value=f"{result['conversion_rate']:.2f}%")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            # ACOS
            cell = self.ws_out.cell(row=current_row, column=start_col + 4, value=f"{result['acos']:.2f}%")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            # 广告花费
            cell = self.ws_out.cell(row=current_row, column=start_col + 5, value=f"{result['ad_spend']:.2f}")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            # 销售额
            cell = self.ws_out.cell(row=current_row, column=start_col + 6, value=f"{result['sales']:.2f}")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
            # cell.fill = BLUE_FILL
            
            current_row += 1

    def _write_sales_table_row(self, row, start_col, rank, item, fill_color):
        """写入销量表格行"""
        sku = item.get("sku", "")
        u1 = safe_round_units(item.get("u1", 0))
        u0 = safe_round_units(item.get("u0", 0))
        growth_rate = item.get("growth_rate", 0)
        
        # 排名
        cell = self.ws_out.cell(row=row, column=start_col, value=rank)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # SKU
        cell = self.ws_out.cell(row=row, column=start_col + 1, value=sku)
        cell.border = THIN_BORDER
        cell.fill = fill_color
        
        # 上周销量
        cell = self.ws_out.cell(row=row, column=start_col + 2, value=u1)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # 本周销量
        cell = self.ws_out.cell(row=row, column=start_col + 3, value=u0)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # 环比增长
        cell = self.ws_out.cell(row=row, column=start_col + 4, value=format_percent_with_sign(growth_rate))
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color

    def _write_conversion_table_row(self, row, start_col, rank, item, fill_color):
        """写入转化率表格行"""
        sku = item.get("sku", "")
        p1 = item.get("p1", 0) * 100  # 转为百分比
        p0 = item.get("p0", 0) * 100  # 转为百分比
        growth_rate = item.get("growth_rate", 0)
        
        # 排名
        cell = self.ws_out.cell(row=row, column=start_col, value=rank)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # SKU
        cell = self.ws_out.cell(row=row, column=start_col + 1, value=sku)
        cell.border = THIN_BORDER
        cell.fill = fill_color
        
        # 上周转化率
        cell = self.ws_out.cell(row=row, column=start_col + 2, value=f"{p1:.2f}%")
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # 本周转化率
        cell = self.ws_out.cell(row=row, column=start_col + 3, value=f"{p0:.2f}%")
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color
        
        # 环比增长
        cell = self.ws_out.cell(row=row, column=start_col + 4, value=format_percent_with_sign(growth_rate))
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = fill_color

    def _adjust_table_column_width(self, start_col, end_col, start_row, end_row):
        """调整表格列宽"""
        for col in range(start_col, end_col + 1):
            max_length = 0
            for row in range(start_row, end_row + 1):
                cell = self.ws_out.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # 设置列宽，加一些余量
            adjusted_width = min(max_length + 2, 50)
            column_letter = openpyxl.utils.get_column_letter(col)
            self.ws_out.column_dimensions[column_letter].width = adjusted_width

    def write_analysis_table(self, top2_sales, bottom2_sales, top2_conv, bottom2_conv):
        """写入销量和转化率的分析表格（双行标题版本）"""
        self._write_sales_analysis_table(top2_sales, bottom2_sales)
        self._write_conversion_analysis_table(top2_conv, bottom2_conv)

    def _write_sales_analysis_table(self, top2_sales, bottom2_sales):
        """写入销量分析表格（双行标题）"""
        start_col = SALES_TABLE_START_COL
        start_row = SALES_TABLE_START_ROW
        
        # 大标题行 - "销量分析"
        title_cell = self.ws_out.cell(row=start_row, column=start_col, value="销量分析（订单数量>14的SKU）")
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = THIN_BORDER
        self.ws_out.merge_cells(start_row=start_row, start_column=start_col, 
                            end_row=start_row, end_column=start_col + 4)
        
        # 表头行
        header_row = start_row + 1
        headers = ["排名", "SKU", "上周销量", "本周销量", "环比增长"]
        for col_idx, header in enumerate(headers):
            cell = self.ws_out.cell(row=header_row, column=start_col + col_idx, value=header)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
        
        # 写入增长SKU（红色）
        current_row = header_row + 1
        for i, item in enumerate(top2_sales, 1):
            rank_label = f"Top{i}"  
            self._write_sales_table_row(current_row, start_col, rank_label, item, RED_FILL)
            current_row += 1
        
        # 写入下降SKU（绿色）  
        for i, item in enumerate(bottom2_sales, 1):
            rank_label = f"Bottom{i}"  
            self._write_sales_table_row(current_row, start_col, rank_label, item, GREEN_FILL)
            current_row += 1
            
        # 调整列宽
        self._adjust_table_column_width(start_col, start_col + len(headers) - 1, start_row, current_row - 1)

        # 新增：重新设置第一列为固定宽度
        first_col_letter = openpyxl.utils.get_column_letter(start_col)
        self.ws_out.column_dimensions[first_col_letter].width = 8  # 排名列固定为8字符

    def _write_conversion_analysis_table(self, top2_conv, bottom2_conv):
        """写入转化率分析表格（双行标题）"""
        start_col = CONV_TABLE_START_COL
        start_row = CONV_TABLE_START_ROW
        
        # 大标题行 - "转化率分析"
        title_cell = self.ws_out.cell(row=start_row, column=start_col, value="转化率分析（订单数量>14的SKU）")
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.border = THIN_BORDER
        title_cell.alignment = CENTER_ALIGNMENT
        self.ws_out.merge_cells(start_row=start_row, start_column=start_col, 
                            end_row=start_row, end_column=start_col + 4)
        
        # 表头行
        header_row = start_row + 1
        headers = ["排名", "SKU", "上周转化率", "本周转化率", "环比增长"]
        for col_idx, header in enumerate(headers):
            cell = self.ws_out.cell(row=header_row, column=start_col + col_idx, value=header)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT
    
        # 写入增长SKU（红色）
        current_row = header_row + 1
        for i, item in enumerate(top2_conv, 1):
            rank_label = f"Top{i}"  
            self._write_conversion_table_row(current_row, start_col, rank_label, item, RED_FILL)
            current_row += 1
        
        # 写入下降SKU（绿色）
        for i, item in enumerate(bottom2_conv, 1):
            rank_label = f"Bottom{i}"  
            self._write_conversion_table_row(current_row, start_col, rank_label, item, GREEN_FILL)
            current_row += 1


# 五行四列表格
    def write_summary_table(self, file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest):
        """生成五列四行的汇总表"""
        start_col = SUMMARY_START_COL
        start_row = SUMMARY_START_ROW

        col_titles = ["", "上周", "本周", "差额", "环比变化"]
        row_titles = ["销售额", "订单数", "店铺综合ACOS"]

        # 新增：调整第一列宽度为20个字符
        first_col_letter = openpyxl.utils.get_column_letter(start_col)
        self.ws_out.column_dimensions[first_col_letter].width = 20

        # 清空原有区域
        self._clear_range(
            start_row,
            start_col,
            start_row + 1 + len(row_titles) + 1,
            start_col + len(col_titles) + 1
        )
        
                # 新增：写入标题行 "店铺综合分析"
        title_row = start_row - 1
        title_cell = self.ws_out.cell(row=title_row, column=start_col, value="店铺综合分析")
        title_cell.font = Font(bold=True, size=14, color="000000")  # 加粗，14号字体
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = THIN_BORDER
        # 合并标题单元格（跨所有列）
        self.ws_out.merge_cells(
            start_row=title_row, 
            start_column=start_col, 
            end_row=title_row, 
            end_column=start_col + len(col_titles) - 1
        )

        # 写列标题
        for i, title in enumerate(col_titles):
            cell = self.ws_out.cell(row=start_row, column=start_col + i, value=title)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER

        # 写行标题
        for r_idx, rname in enumerate(row_titles):
            row_num = start_row + 1 + r_idx
            cell = self.ws_out.cell(row=row_num, column=start_col, value=rname)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            
            for c in range(1, len(col_titles)):
                cnum = start_col + c
                self.ws_out.cell(row=row_num, column=cnum).alignment = CENTER_ALIGNMENT
                self.ws_out.cell(row=row_num, column=cnum).border = THIN_BORDER

        # 计算并写入具体数据
        self._write_sales_data(file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row)
        self._write_orders_data(file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row)
        self._write_acos_data(file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row)

        # 加样式：给整个 summary 区域加框并居中
        for r in range(start_row, start_row + 1 + len(row_titles)):
            for c in range(start_col, start_col + len(col_titles)):
                cell = self.ws_out.cell(row=r, column=c)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT

    def _write_sales_data(self, file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row):
        """写入销售额数据"""
        from data_loader import sum_column_total
        
        total_ordered_prev2 = sum_column_total(file_prev2, ["ordered product sales", "ordered product", "ordered sales"])
        total_ordered_prev = sum_column_total(file_prev, ["ordered product sales", "ordered product", "ordered sales"])
        total_ordered_latest = sum_column_total(file_latest, ["ordered product sales", "ordered product", "ordered sales"])

        diff_ordered_prev, days_prev_adjust = calc_diff_or_cumulative(
            Decimal(total_ordered_prev), date_prev, Decimal(total_ordered_prev2), date_prev2
        )
        diff_ordered_latest, days_latest_adjust = calc_diff_or_cumulative(
            Decimal(total_ordered_latest), date_latest, Decimal(total_ordered_prev), date_prev
        )

        m1 = adjust_to_7_days(diff_ordered_prev, days_prev_adjust)
        m0 = adjust_to_7_days(diff_ordered_latest, days_latest_adjust)

        if m1 is None or m0 is None:
            print("汇总销售额计算出错（日期跨度异常）")
            return

        # 写销售额行
        sales_row = start_row + 1
        self.ws_out.cell(row=sales_row, column=start_col + 1, value=round(m1, 2))
        self.ws_out.cell(row=sales_row, column=start_col + 2, value=round(m0, 2))
        
        t1_coord = self.ws_out.cell(row=sales_row, column=start_col + 1).coordinate
        t0_coord = self.ws_out.cell(row=sales_row, column=start_col + 2).coordinate
        
        diff_cell = self.ws_out.cell(row=sales_row, column=start_col + 3)
        diff_cell.value = f"={t0_coord}-{t1_coord}"
        diff_cell.border = THIN_BORDER
        diff_cell.alignment = CENTER_ALIGNMENT
        
        ratio_cell = self.ws_out.cell(row=sales_row, column=start_col + 4)
        ratio_cell.value = f'=IF({t1_coord}=0,"0%",IF(({t0_coord}-{t1_coord})/{t1_coord}>0,"+" & TEXT(({t0_coord}-{t1_coord})/{t1_coord},"0.00%"),TEXT(({t0_coord}-{t1_coord})/{t1_coord},"0.00%")))'
        ratio_cell.border = THIN_BORDER
        ratio_cell.alignment = CENTER_ALIGNMENT
        
        self.ws_out.cell(row=sales_row, column=start_col + 1).border = THIN_BORDER
        self.ws_out.cell(row=sales_row, column=start_col + 2).border = THIN_BORDER

    def _write_orders_data(self, file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row):
        """写入订单数据"""
        from data_loader import sum_column_total
        
        total_units_prev2 = sum_column_total(file_prev2, ["units ordered", "unitsordered"])
        total_units_prev = sum_column_total(file_prev, ["units ordered", "unitsordered"])
        total_units_latest = sum_column_total(file_latest, ["units ordered", "unitsordered"])

        diff_units_prev, days_prev_units = calc_diff_or_cumulative(
            Decimal(total_units_prev), date_prev, Decimal(total_units_prev2), date_prev2
        )
        diff_units_latest, days_latest_units = calc_diff_or_cumulative(
            Decimal(total_units_latest), date_latest, Decimal(total_units_prev), date_prev
        )

        u1_val = adjust_to_7_days(diff_units_prev, days_prev_units)
        u0_val = adjust_to_7_days(diff_units_latest, days_latest_units)

        if u1_val is None or u0_val is None:
            print("汇总订单数计算出错（日期跨度异常）")
            return

        orders_row = start_row + 2
        self.ws_out.cell(row=orders_row, column=start_col + 1, value=safe_round_units(u1_val))
        self.ws_out.cell(row=orders_row, column=start_col + 2, value=safe_round_units(u0_val))
        
        t1_coord = self.ws_out.cell(row=orders_row, column=start_col + 1).coordinate
        t0_coord = self.ws_out.cell(row=orders_row, column=start_col + 2).coordinate
        
        self.ws_out.cell(row=orders_row, column=start_col + 3, value=f"={t0_coord}-{t1_coord}").border = THIN_BORDER
        self.ws_out.cell(row=orders_row, column=start_col + 4, value=f'=IF({t1_coord}=0,"0%",IF(({t0_coord}-{t1_coord})/{t1_coord}>0,"+" & TEXT(({t0_coord}-{t1_coord})/{t1_coord},"0.00%"),TEXT(({t0_coord}-{t1_coord})/{t1_coord},"0.00%")))').border = THIN_BORDER
        
        self.ws_out.cell(row=orders_row, column=start_col + 1).alignment = CENTER_ALIGNMENT
        self.ws_out.cell(row=orders_row, column=start_col + 1).border = THIN_BORDER
        self.ws_out.cell(row=orders_row, column=start_col + 2).alignment = CENTER_ALIGNMENT
        self.ws_out.cell(row=orders_row, column=start_col + 2).border = THIN_BORDER

    def _write_acos_data(self, file_prev2, file_prev, file_latest, date_prev2, date_prev, date_latest, start_col, start_row):
        """写入ACOS数据"""
        from data_loader import sum_column_total
        
        total_ad_prev2 = sum_column_total(file_prev2, ["ad-total-spend", "ad total spend", "ad_total_spend", "ad spend"])
        total_ad_prev = sum_column_total(file_prev, ["ad-total-spend", "ad total spend", "ad_total_spend", "ad spend"])
        total_ad_latest = sum_column_total(file_latest, ["ad-total-spend", "ad total spend", "ad_total_spend", "ad spend"])

        diff_ad_prev, days_prev_ad = calc_diff_or_cumulative(
            Decimal(total_ad_prev), date_prev, Decimal(total_ad_prev2), date_prev2
        )
        diff_ad_latest, days_latest_ad = calc_diff_or_cumulative(
            Decimal(total_ad_latest), date_latest, Decimal(total_ad_prev), date_prev
        )

        a1 = adjust_to_7_days(diff_ad_prev, days_prev_ad)
        a0 = adjust_to_7_days(diff_ad_latest, days_latest_ad)

        # 重新计算销售额用于ACOS计算
        total_ordered_prev2 = sum_column_total(file_prev2, ["ordered product sales", "ordered product", "ordered sales"])
        total_ordered_prev = sum_column_total(file_prev, ["ordered product sales", "ordered product", "ordered sales"])
        total_ordered_latest = sum_column_total(file_latest, ["ordered product sales", "ordered product", "ordered sales"])

        diff_ordered_prev, days_prev_adjust = calc_diff_or_cumulative(
            Decimal(total_ordered_prev), date_prev, Decimal(total_ordered_prev2), date_prev2
        )
        diff_ordered_latest, days_latest_adjust = calc_diff_or_cumulative(
            Decimal(total_ordered_latest), date_latest, Decimal(total_ordered_prev), date_prev
        )

        m1 = adjust_to_7_days(diff_ordered_prev, days_prev_adjust)
        m0 = adjust_to_7_days(diff_ordered_latest, days_latest_adjust)

        acos_row = start_row + 3
        # 防止除以 0
        if m1 == 0:
            self.ws_out.cell(row=acos_row, column=start_col + 1, value="0.00%")
        else:
            self.ws_out.cell(row=acos_row, column=start_col + 1, value=f"{(a1/m1 if m1 else 0)*100:.2f}%")
        if m0 == 0:
            self.ws_out.cell(row=acos_row, column=start_col + 2, value="0.00%")
        else:
            self.ws_out.cell(row=acos_row, column=start_col + 2, value=f"{(a0/m0 if m0 else 0)*100:.2f}%")

        t1_coord = self.ws_out.cell(row=acos_row, column=start_col + 1).coordinate
        t0_coord = self.ws_out.cell(row=acos_row, column=start_col + 2).coordinate
        
        self.ws_out.cell(row=acos_row, column=start_col + 3).value = f'=IF({t1_coord}="0.00%","0%",IF((VALUE({t0_coord})-VALUE({t1_coord}))>0,"+" & TEXT(VALUE({t0_coord})-VALUE({t1_coord}),"0.00%"),TEXT(VALUE({t0_coord})-VALUE({t1_coord}),"0.00%")))'
        self.ws_out.cell(row=acos_row, column=start_col + 4).value = f'=IF({t1_coord}="0.00%","0%",IF((VALUE({t0_coord})-VALUE({t1_coord}))/IF(VALUE({t1_coord})=0,1,VALUE({t1_coord}))>0,"+" & TEXT((VALUE({t0_coord})-VALUE({t1_coord}))/IF(VALUE({t1_coord})=0,1,VALUE({t1_coord})),"0.00%"),TEXT((VALUE({t0_coord})-VALUE({t1_coord}))/IF(VALUE({t1_coord})=0,1,VALUE({t1_coord})),"0.00%")))'

    def write_sum_row(self):
        """
        在原始数据表最后一行添加求和行
        求和列：Sessions, Page Views, Units Ordered, Ordered Product Sales, AD-Total-Spend, AD-Total-Sale
        平均列：Unit Session Percentage(%)
        计算列：Total-acos(%) = AD-Total-Spend和 / Ordered Product Sales和
                AD-ACOS(%) = AD-Total-Spend和 / AD-Total-Sale和
        """
        # 定义列配置：(关键字, 计算类型, 列标识)
        # 计算类型: 'sum'=求和, 'avg'=平均, 'total_acos'=总ACOS计算, 'ad_acos'=广告ACOS计算,'ad_sales_ratio'=广告销售额占比
        column_config = [
            (["sessions"], 'sum', 'sessions'),
            (["page views", "pageviews", "page-views"], 'sum', 'page_views'),
            (["unit session percentage", "unit session percentage(%)", "unit session %"], 'avg', 'unit_session_pct'),
            (["units ordered", "unitsordered"], 'sum', 'units'),
            (["ordered product sales", "ordered product", "ordered sales", "orderedproduct"], 'sum', 'ordered_sales'),
            (["total-acos(%)", "total acos(%)", "total acos", "total-acos"], 'total_acos', 'total_acos'),
            (["ad-total-spend", "ad total spend", "ad_total_spend", "ad spend"], 'sum', 'ad_spend'),
            (["ad-acos(%)", "ad acos(%)", "ad acos", "ad-acos"], 'ad_acos', 'ad_acos'),
            (["广告销售占比(%)", "广告销售占比"], 'ad_sales_ratio', 'ad_sales_ratio'),
            (["ad-total-sale", "ad total sale", "ad_total_sale", "ad sales"], 'sum', 'ad_sales')
        ]

        # 找到各列的索引和计算类型
        col_info = []  # [(col_idx, calc_type, col_id), ...]
        for keywords, calc_type, col_id in column_config:
            col_idx = None
            for idx, cell in enumerate(self.ws_out[1], start=1):
                if cell.value and any(kw.lower() in str(cell.value).lower() for kw in keywords):
                    col_idx = idx
                    break
            if col_idx:
                col_info.append((col_idx, calc_type, col_id))

        # 确定数据最后一行（跳过空行）
        max_data_row = self.ws_out.max_row
        while max_data_row > 1:
            has_data = False
            for col_idx, _, _ in col_info:
                cell = self.ws_out.cell(row=max_data_row, column=col_idx)
                if cell.value is not None:
                    has_data = True
                    break
            if has_data:
                break
            max_data_row -= 1

        # 如果数据最后一行已经是"总计"行，先删除
        last_cell_value = self.ws_out.cell(row=max_data_row, column=1).value
        if last_cell_value and "总计" in str(last_cell_value):
            self.ws_out.delete_rows(max_data_row)
            max_data_row -= 1

        # 在数据最后一行下方添加总计行
        sum_row = max_data_row + 1

        # 写入"总计"标签（第一列）
        first_col = 1
        label_cell = self.ws_out.cell(row=sum_row, column=first_col, value="总计")
        label_cell.font = BOLD_FONT
        label_cell.alignment = CENTER_ALIGNMENT

        # 存储各列的SUM公式坐标
        sum_formulas = {}  # {col_id: cell_coordinate, ...}

        # 第一步：处理所有 sum 类型，先存储坐标
        for col_idx, calc_type, col_id in col_info:
            if calc_type == 'sum':
                cell = self.ws_out.cell(row=sum_row, column=col_idx)
                cell.font = BOLD_FONT
                cell.alignment = CENTER_ALIGNMENT
                start_cell = self.ws_out.cell(row=2, column=col_idx).coordinate
                end_cell = self.ws_out.cell(row=max_data_row, column=col_idx).coordinate
                cell.value = f"=SUM({start_cell}:{end_cell})"
                sum_formulas[col_id] = cell.coordinate

        # 第二步：处理其他类型（avg, total_acos, ad_acos）
        for col_idx, calc_type, col_id in col_info:
            if calc_type == 'sum':
                continue  # 已在第一步处理

            cell = self.ws_out.cell(row=sum_row, column=col_idx)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGNMENT

            start_cell = self.ws_out.cell(row=2, column=col_idx).coordinate
            end_cell = self.ws_out.cell(row=max_data_row, column=col_idx).coordinate

            if calc_type == 'avg':
                # 求平均（保留两位小数）
                cell.value = f"=AVERAGE({start_cell}:{end_cell})"
                cell.number_format = '0.00'
            elif calc_type == 'total_acos':
                # Total-acos(%) = AD-Total-Spend和 / Ordered Product Sales和（保留两位小数）
                ad_spend_coord = sum_formulas.get('ad_spend')
                ordered_sales_coord = sum_formulas.get('ordered_sales')
                if ad_spend_coord and ordered_sales_coord:
                    cell.value = f"=IF({ordered_sales_coord}=0,0,{ad_spend_coord}/{ordered_sales_coord}*100)"
                    cell.number_format = '0.00'
            elif calc_type == 'ad_acos':
                # AD-ACOS(%) = AD-Total-Spend和 / AD-Total-Sale和（保留两位小数）
                ad_spend_coord = sum_formulas.get('ad_spend')
                ad_sales_coord = sum_formulas.get('ad_sales')
                if ad_spend_coord and ad_sales_coord:
                    cell.value = f"=IF({ad_sales_coord}=0,0,{ad_spend_coord}/{ad_sales_coord}*100)"
                    cell.number_format = '0.00'
            elif calc_type == 'ad_sales_ratio':
                # 广告销售占比(%) = AD-Total-Sale和 / Ordered Product Sales和（保留两位小数）
                ad_sales_coord = sum_formulas.get('ad_sales')
                ordered_sales_coord = sum_formulas.get('ordered_sales')
                if ad_sales_coord and ordered_sales_coord:
                    cell.value = f"=IF({ordered_sales_coord}=0,0,{ad_sales_coord}/{ordered_sales_coord}*100)"
                    cell.number_format = '0.00'
