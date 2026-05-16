# data_loader.py
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from utils import *

def load_sheet_data(file_path):
    """
    加载 Excel 第一个 sheet，
    返回： (ws, headers_map, data_dict)
    data_dict: { sku: {"sessions": number, "page_views": number, "units": number, "ordered_sales": number, "ad_spend": number, "ad_sales": number, "acos": number} }
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.worksheets[0]
    
    # 构建 headers map (col_name -> index)
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        headers[idx] = str(cell.value).strip()
    
    # 找列索引：使用关键字包含匹配
    def find_col_by_keywords(keywords):
        for idx, title in headers.items():
            low = title.lower()
            for kw in keywords:
                if kw.lower() in low:
                    return idx
        return None

    sku_col = find_col_by_keywords(["sku"])
    sessions_col = find_col_by_keywords(["sessions"])
    page_views_col = find_col_by_keywords(["page views", "pageviews", "page-views"])
    units_col = find_col_by_keywords(["units ordered", "unitsordered", "unitsordered"])
    ordered_sales_col = find_col_by_keywords(["ordered product sales", "ordered product", "ordered sales", "orderedproduct"])
    ad_spend_col = find_col_by_keywords(["ad-total-spend", "ad total spend", "ad_total_spend", "ad total", "ad-spend"])
    ad_sales_col = find_col_by_keywords(["ad-total-sale", "ad total sale", "ad_total_sale", "ad sales"])

    data = {}
    # iterate rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not sku_col:
            break
        sku = row[sku_col - 1]
        if sku is None:
            continue
            
        sessions = 0
        page_views = 0
        units = 0
        ordered_sales = 0
        ad_spend = 0
        ad_sales = 0

        if sessions_col:
            v = row[sessions_col - 1]
            sessions = v if isinstance(v, (int, float)) else 0
        if page_views_col:
            v = row[page_views_col - 1]
            page_views = v if isinstance(v, (int, float)) else 0
        if units_col:
            v = row[units_col - 1]
            units = v if isinstance(v, (int, float)) else 0
        if ordered_sales_col:
            v = row[ordered_sales_col - 1]
            ordered_sales = v if isinstance(v, (int, float)) else 0
        if ad_spend_col:
            v = row[ad_spend_col - 1]
            ad_spend = v if isinstance(v, (int, float)) else 0
        if ad_sales_col:
            v = row[ad_sales_col - 1]
            ad_sales = v if isinstance(v, (int, float)) else 0

        # 计算 ACOS (%)
        acos = None
        if ad_sales > 0:
            acos = (ad_spend / ad_sales) * 100

        data[str(sku)] = {
            "sessions": sessions,
            "page_views": page_views,
            "units": units,
            "ordered_sales": ordered_sales,
            "ad_spend": ad_spend,
            "ad_sales": ad_sales,
            "acos": acos
        }
    return ws, headers, data

def sum_column_total(file_path, keywords):
    """对 file_path 第一个 sheet 中匹配 keywords 的列求和（整列数值），返回总和（保留两位小数）"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.worksheets[0]

    # 找列索引
    col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and any(kw.lower() in str(cell.value).lower() for kw in keywords):
            col_idx = idx
            break
    if not col_idx:
        return Decimal('0.00')

    total = Decimal('0.00')
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        v = row[0]
        if isinstance(v, (int, float)):
            total += Decimal(str(v))

    # 保留两位小数
    total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return total

def calculate_weekly_data(data_prev2, data_prev, data_latest, date_prev2, date_prev, date_latest):
    """计算周度数据"""
    days_prev = days_between(date_prev2, date_prev)
    days_latest = days_between(date_prev, date_latest)
    
    if days_prev <= 0 or days_latest <= 0:
        print("日期解析有问题，两个相邻文件日期必须不同。")
        return None, None

    week1 = {}  # 上周数据
    week0 = {}  # 本周数据

    # 上周 (prev vs prev2)
    for sku, v in data_prev.items():
        if sku in data_prev2:
            prev_val_sess = Decimal(str(data_prev[sku]["sessions"]))
            prev2_val_sess = Decimal(str(data_prev2[sku]["sessions"]))
            diff_sessions, days_sess = calc_diff_or_cumulative(prev_val_sess, date_prev, prev2_val_sess, date_prev2)

            prev_val_units = Decimal(str(data_prev[sku]["units"]))
            prev2_val_units = Decimal(str(data_prev2[sku]["units"]))
            diff_units, days_units = calc_diff_or_cumulative(prev_val_units, date_prev, prev2_val_units, date_prev2)

            # 新增：广告花费和销售额计算
            prev_val_ad = Decimal(str(data_prev[sku]["ad_spend"]))
            prev2_val_ad = Decimal(str(data_prev2[sku]["ad_spend"]))
            diff_ad, days_ad = calc_diff_or_cumulative(prev_val_ad, date_prev, prev2_val_ad, date_prev2)

            prev_val_adsales = Decimal(str(data_prev[sku]["ad_sales"]))
            prev2_val_adsales = Decimal(str(data_prev2[sku]["ad_sales"]))
            diff_adsales, days_adsales = calc_diff_or_cumulative(prev_val_adsales, date_prev, prev2_val_adsales, date_prev2)

            prev_val_sales = Decimal(str(data_prev[sku]["ordered_sales"]))
            prev2_val_sales = Decimal(str(data_prev2[sku]["ordered_sales"]))
            diff_sales, days_sales = calc_diff_or_cumulative(prev_val_sales, date_prev, prev2_val_sales, date_prev2)

            adj_sessions = adjust_to_7_days(diff_sessions, days_sess)
            adj_units = adjust_to_7_days(diff_units, days_units)
            #新增
            adj_ad = adjust_to_7_days(diff_ad, days_ad)
            adj_sales = adjust_to_7_days(diff_sales, days_sales)
            adj_adsales = adjust_to_7_days(diff_adsales, days_adsales)

            if adj_sessions is not None and adj_units is not None:
                week1[sku] = {"sessions": adj_sessions, "units": adj_units, "ad_spend": adj_ad, "ordered_sales": adj_sales, "ad_sales":adj_adsales }

    # 本周 (latest vs prev)
    for sku, v in data_latest.items():
        if sku in data_prev:
            latest_val_sess = Decimal(str(data_latest[sku]["sessions"]))
            prev_val_sess = Decimal(str(data_prev[sku]["sessions"]))
            diff_sessions, days_sess = calc_diff_or_cumulative(latest_val_sess, date_latest, prev_val_sess, date_prev)

            latest_val_units = Decimal(str(data_latest[sku]["units"]))
            prev_val_units = Decimal(str(data_prev[sku]["units"]))
            diff_units, days_units = calc_diff_or_cumulative(latest_val_units, date_latest, prev_val_units, date_prev)

            # 新增：广告花费和销售额计算
            latest_val_ad = Decimal(str(data_latest[sku]["ad_spend"]))
            prev_val_ad = Decimal(str(data_prev[sku]["ad_spend"]))
            diff_ad, days_ad = calc_diff_or_cumulative(latest_val_ad, date_latest, prev_val_ad, date_prev)

            latest_val_sales = Decimal(str(data_latest[sku]["ordered_sales"]))
            prev_val_sales = Decimal(str(data_prev[sku]["ordered_sales"]))
            diff_sales, days_sales = calc_diff_or_cumulative(latest_val_sales, date_latest, prev_val_sales, date_prev)

            latest_val_adsales = Decimal(str(data_latest[sku]["ad_sales"]))
            prev_val_adsales = Decimal(str(data_prev[sku]["ad_sales"]))
            diff_adsales, days_adsales = calc_diff_or_cumulative(latest_val_adsales, date_latest, prev_val_adsales, date_prev)

            adj_sessions = adjust_to_7_days(diff_sessions, days_sess)
            adj_units = adjust_to_7_days(diff_units, days_units)
            #新增
            adj_ad = adjust_to_7_days(diff_ad, days_ad)
            adj_sales = adjust_to_7_days(diff_sales, days_sales)
            adj_adsales = adjust_to_7_days(diff_adsales, days_adsales)

            if adj_sessions is not None and adj_units is not None:
                week0[sku] = {"sessions": adj_sessions, "units": adj_units, "ad_spend": adj_ad, "ordered_sales": adj_sales,"ad_sales":adj_adsales }

    return week1, week0

#增加acos分析
def calculate_acos_analysis(week1, week0):
    """
    计算ACOS分析结果：找出本周ACOS>10%且销量增长最快的SKU
    返回: 销量增长最快的前3个SKU列表（本周ACOS>10%）
    """
    acos_results = []
    
    # 遍历所有在week1和week0中都存在的SKU
    common_skus = set(week1.keys()) & set(week0.keys())
    
    for sku in common_skus:
        try:
            # 从已经7天标准化后的数据中获取广告花费和销售额
            a1 = week1[sku]["ad_spend"]  # 上周广告花费（已7天标准化）
            a2 = week0[sku]["ad_spend"]  # 本周广告花费（已7天标准化）
            t1 = week1[sku]["ad_sales"]  # 上周广告销售额（已7天标准化）
            t2 = week0[sku]["ad_sales"]  # 本周广告销售额（已7天标准化）

            # 获取总销售额数据用于计算广告出单占比
            # total_sales1 = week1[sku]["ordered_sales"]  # 上周总销售额
            total_sales2 = week0[sku]["ordered_sales"]  # 本周总销售额
            
            # 获取销量数据用于增长计算
            u1 = week1[sku]["units"]  # 上周销量
            u2 = week0[sku]["units"]  # 本周销量
            
            # 检查数据有效性
            if (a1 is None or a2 is None or t1 is None or t2 is None or 
                u1 is None or u2 is None or t2 <= 0):  # 本周销售额必须为正数
                continue
            
            # 计算本周ACOS
            acos2 = (a2 / t2) * 100
            
            # 只保留本周ACOS大于10%的SKU
            if acos2 <= 10.0:
                continue
            
            # 计算销量增长
            if u1 > 0:  # 避免除零
                sales_growth_rate = ((u2 - u1) / u1) * 100
            else:
                sales_growth_rate = float('inf')  # 如果上周销量为0，本周有销量，视为无限增长

            # 计算广告出单占比（广告销售额 / 总销售额 × 100%）
            ad_sales_ratio2 = (t2 / total_sales2) * 100  # 本周广告出单占比
            # ad_sales_ratio1 = (t1 / total_sales1) * 100 if total_sales1 > 0 else 0  # 上周广告出单占比
                
            acos_results.append({
                'sku': sku,
                'acos1': float((a1 / t1) * 100) if t1 > 0 else 0,  # 上周ACOS
                'acos2': float(acos2),  # 本周ACOS
                'sales_growth': float(u2 - u1),  # 销量增长绝对值
                'sales_growth_rate': float(sales_growth_rate),  # 销量增长率
                'u1': float(u1),  # 上周销量
                'u2': float(u2),   # 本周销量
                # 'ad_sales_ratio1': float(ad_sales_ratio1),  # 上周广告出单占比
                'ad_sales_ratio2': float(ad_sales_ratio2)  # 本周广告出单占比
            })
                
        except (ZeroDivisionError, KeyError, TypeError, ValueError) as e:
            # 跳过计算错误的SKU
            continue
    
    # 在排序之前打印所有符合条件的SKU
    # print("\n📋 所有符合条件的SKU（本周ACOS>10%）：")
    if not acos_results:
        print("   ❌ 没有找到符合条件的SKU")
    else:
        print(f"   共找到 {len(acos_results)} 个SKU")
        # 按销量增长率排序显示所有符合条件的SKU
        # all_results_sorted = sorted(acos_results, key=lambda x: x['sales_growth_rate'], reverse=True)
        # for i, result in enumerate(all_results_sorted, 1):
        #     print(f"   {i:2d}. {result['sku']:<15} "
        #           f"ACOS: {result['acos2']:5.1f}%, "
        #           f"增长: {result['sales_growth_rate']:6.1f}%, "
        #           f"销量: {result['u1']:3.0f}→{result['u2']:3.0f},"
        #           f"广告占比: {result['ad_sales_ratio2']:5.1f}%")
    # 按销量增长率降序排列，取前3个
    top_acos_results = sorted(acos_results, key=lambda x: x['sales_growth_rate'], reverse=True)[:3]
    return top_acos_results

#增加
def analyze_high_conversion_products(week1, week0, data_latest):
    """
    分析7天订单数量大于10,转化率最高的3个产品
    返回: 转化率最高的前3个SKU列表
    """
    results = []
    
    # 遍历所有在week1和week0中都存在的SKU
    common_skus = set(week1.keys()) & set(week0.keys())
    
    for sku in common_skus:
        try:
            # 获取销量和会话数据
            u1 = week1[sku]["units"]  # 上周销量
            u2 = week0[sku]["units"]  # 本周销量
            s1 = week1[sku]["sessions"]  # 上周会话
            s2 = week0[sku]["sessions"]  # 本周会话
            
            # 检查订单数量条件：本周订单数量大于10
            if u2 <= 10:
                continue
                
            # 检查会话量有效性
            if s2 <= 0:
                continue
                
            # 计算本周转化率
            conversion_rate = (u2 / s2) * 100
            
            results.append({
                'sku': sku,
                'units': float(u2),  # 本周销量
                'sessions': float(s2),  # 本周会话
                'conversion_rate': float(conversion_rate),  # 转化率
                'last_week_units': float(u1),  # 上周销量（用于对比）
                'last_week_conversion': float((u1 / s1) * 100) if s1 > 0 else 0  # 上周转化率
            })
                
        except (ZeroDivisionError, KeyError, TypeError, ValueError) as e:
            continue
    
    # 按转化率降序排列，取前3个
    top_results = sorted(results, key=lambda x: x['conversion_rate'], reverse=True)[:3]
    return top_results

#增加
def analyze_high_conversion_low_acos_products(week1, week0, data_latest):
    """
    分析7天订单数量大于10,且转化率大于10%,且ACOS小于5%的产品
    返回: 符合条件的SKU列表
    """
    results = []
    
    # 遍历所有在week1和week0中都存在的SKU
    common_skus = set(week1.keys()) & set(week0.keys())
    
    for sku in common_skus:
        try:
            # 获取销量、会话和广告数据
            u1 = week1[sku]["units"]  # 上周销量
            u2 = week0[sku]["units"]  # 本周销量
            s1 = week1[sku]["sessions"]  # 上周会话
            s2 = week0[sku]["sessions"]  # 本周会话
            a2 = week0[sku]["ad_spend"]  # 本周广告花费
            t2 = week0[sku]["ordered_sales"]  # 本周销售额
            
            # 检查订单数量条件：本周订单数量大于10
            if u2 <= 10:
                continue
                
            # 检查会话量有效性
            if s2 <= 0:
                continue
                
            # 计算本周转化率
            conversion_rate = (u2 / s2) * 100
            
            # 检查转化率条件：转化率大于10%
            if conversion_rate <= 10.0:
                continue
                
            # 计算本周ACOS
            if t2 > 0:
                acos = (a2 / t2) * 100
            else:
                acos = 0
                
            # 检查ACOS条件：ACOS小于5%
            if acos >= 5.0:
                continue
                
            results.append({
                'sku': sku,
                'units': float(u2),  # 本周销量
                'sessions': float(s2),  # 本周会话
                'conversion_rate': float(conversion_rate),  # 转化率
                'acos': float(acos),  # ACOS
                'ad_spend': float(a2),  # 广告花费
                'sales': float(t2)  # 销售额
            })
                
        except (ZeroDivisionError, KeyError, TypeError, ValueError) as e:
            continue
    
    # 可以按转化率或其他指标排序
    sorted_results = sorted(results, key=lambda x: x['conversion_rate'], reverse=True)
    return sorted_results

def analyze_sku_performance(week1, week0):
    """分析SKU性能，返回销量和转化率的top/bottom结果"""
    skus = set(week1.keys()) & set(week0.keys())

    sales_results = []  # {'sku', 'u1', 'u0', 'growth_rate'}
    conv_results = []   # {'sku', 'p1', 'p0', 'growth_rate'}

    for sku in skus:
        u1 = week1[sku]["units"]
        u0 = week0[sku]["units"]
        s1 = week1[sku]["sessions"]
        s0 = week0[sku]["sessions"]

        # 排除周销量 < 14 或点击量 <= 0
        if u1 < 14 or u0 < 14 or s1 <= 0 or s0 <= 0:
            continue

        # 销量环比增长率
        growth_sales = (u0 - u1) / u1 * 100

        # 转化率
        p1 = u1 / s1
        p0 = u0 / s0
        if p1 == 0:
            continue
        growth_conv = (p0 - p1) / p1 * 100

        sales_results.append({"sku": sku, "u1": u1, "u0": u0, "growth_rate": growth_sales})
        conv_results.append({"sku": sku, "p1": p1, "p0": p0, "growth_rate": growth_conv})

    # 动态调整 top/bottom 数量
    def get_top_bottom_dynamic(results, max_count=2):
        """动态获取top和bottom，根据结果数量调整"""
        if not results:
            return [], []
        
        sorted_by = sorted(results, key=lambda x: x["growth_rate"])
        total_count = len(sorted_by)
        
        # 根据总数动态调整显示数量
        if total_count == 0:
            return [], []
        elif total_count == 1:
            # 只有1个SKU，显示为top1
            return [sorted_by[0]], []
        elif total_count == 2:
            # 只有2个SKU，显示top1和bottom1
            return [sorted_by[1]], [sorted_by[0]]
        elif total_count == 3:
            # 有3个：显示top2和bottom1
            return [sorted_by[2], sorted_by[1]], [sorted_by[0]]
        else:
            # 4个或以上，显示top2和bottom2
            bottom_count = min(max_count, total_count // 2)
            top_count = min(max_count, total_count // 2)
            
            bottom = sorted_by[:bottom_count]
            top = sorted_by[-top_count:][::-1]  # highest first
            return top, bottom

    top2_sales, bottom2_sales = get_top_bottom_dynamic(sales_results)
    top2_conv, bottom2_conv = get_top_bottom_dynamic(conv_results)

    return top2_sales, bottom2_sales, top2_conv, bottom2_conv, sales_results, conv_results