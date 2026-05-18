# sku_manager.py - 优化版本
import openpyxl
import os
from openpyxl.comments import Comment
from utils import parse_comment_week_date, parse_brand_country
from config import BLUE_FILL
from openpyxl.styles import PatternFill

def load_or_create_sku_list(sku_dir, brand, country):
    """打开或创建 SKU LIST"""
    filename = f"{brand}-{country}-SKU-LIST.xlsx"
    filepath = os.path.join(sku_dir, filename)
    
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SKU-LIST"
        ws.cell(row=1, column=1, value="SKU")
    
    # 构建SKU映射
    sku_map = {}
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value:
            sku = str(cell.value).strip()
            week_num, comment_date = parse_comment_week_date(cell.comment.text if cell.comment else None)
            sku_map[sku] = {
                'row': cell.row,
                'week_num': week_num,
                'comment_date': comment_date
            }
    
    return wb, ws, sku_map, filepath

def calculate_week_difference(date1, date2):
    """计算两个日期相差的周数（向上取整）"""
    if date1 == date2:
        return 0
    
    days_diff = (date2 - date1).days
    if days_diff <= 0:
        return 0
    
    # 使用整数除法向上取整
    return (days_diff + 6) // 7

def update_sku_new_arrival_status(ws_out, sku_row_map, sku_col_idx, brand, country, sku_dir, current_date, max_weeks=8):
    """更新新上架SKU状态 - 优化版本"""
    wb_sku, ws_sku, sku_list_map, sku_filepath = load_or_create_sku_list(sku_dir, brand, country)
    normalized_sku_row_map = {
        str(sku).strip(): row
        for sku, row in sku_row_map.items()
        if str(sku).strip()
    }
    
    # 处理规则1和规则2：已有标记的SKU
    for sku, sku_info in list(sku_list_map.items()):
        week_num = sku_info['week_num']
        comment_date = sku_info['comment_date']
        
        if week_num is None:
            continue  # 没有批注的SKU，跳过
            
        # 检查当前报表中是否有该SKU
        in_current_report = sku in normalized_sku_row_map
        
        if in_current_report:
            # 规则1：SKU在当前报表中出现
            if comment_date:
                weeks_diff = calculate_week_difference(comment_date, current_date)
                new_week_num = week_num + weeks_diff
            else:
                new_week_num = week_num + 1
        else:
            # 规则2：SKU在当前报表中未出现
            new_week_num = week_num + 1
        
        # 检查是否超过观察期
        if new_week_num > max_weeks:
            # 规则4：超过观察期，移除批注
            ws_sku.cell(row=sku_info['row'], column=1).comment = None
            if in_current_report:
                out_cell = ws_out.cell(row=normalized_sku_row_map[sku], column=sku_col_idx)
                out_cell.comment = None
                out_cell.fill = PatternFill()  # 创建一个空Fill对象
        else:
            # 更新SKU表批注
            new_comment = f"新上架(第{new_week_num}周)@{current_date.isoformat()}"
            ws_sku.cell(row=sku_info['row'], column=1).comment = Comment(new_comment, "Auto")
            
            # 如果在当前报表中，也更新批注
            if in_current_report:
                out_cell = ws_out.cell(row=normalized_sku_row_map[sku], column=sku_col_idx)
                out_cell.comment = Comment(f"新上架(第{new_week_num}周)", "Auto")
                out_cell.fill = BLUE_FILL
    
    # 处理规则3：新SKU首次出现
    for sku, row_num in normalized_sku_row_map.items():
        if sku not in sku_list_map:
            # 规则3：新SKU首次出现
            out_cell = ws_out.cell(row=row_num, column=sku_col_idx)
            out_cell.comment = Comment("新上架(第1周)", "Auto")
            out_cell.fill = BLUE_FILL
            
            # 添加到SKU列表
            ws_sku.append([sku])
            new_row = ws_sku.max_row
            new_comment = f"新上架(第1周)@{current_date.isoformat()}"
            ws_sku.cell(row=new_row, column=1).comment = Comment(new_comment, "Auto")
    
    wb_sku.save(sku_filepath)
    print(f"SKU列表已更新: {sku_filepath}")

def manage_new_arrival_skus(ws_out, sku_row_map, sku_col_idx, file_latest_name, date_latest, base_dir="."):
    """管理新上架SKU的主函数"""
    brand, country = parse_brand_country(file_latest_name)
    if not brand or not country:
        print("无法从文件名解析品牌和国家信息")
        return
    
    # 确保SKU目录存在
    sku_dir = os.path.join(base_dir, "SKU")
    if not os.path.exists(sku_dir):
        os.makedirs(sku_dir)
        print(f"创建SKU目录: {sku_dir}")
    
    update_sku_new_arrival_status(
        ws_out, sku_row_map, sku_col_idx, brand, country, 
        sku_dir, date_latest, max_weeks=8
    )
